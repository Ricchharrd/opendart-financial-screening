from __future__ import annotations

import argparse
import io
import json
import os
import re
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


CORPCODE_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
FS_URL = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
REPORT_CODE = "11011"
PRIMARY_FS_DIV = "CFS"
FALLBACK_FS_DIV = "OFS"
BUILT_IN_API_KEY = "7e0d29c7739a134926f34f371ce16ca2a78e4b99"


METRIC_SPECS = {
    "revenue": {
        "label": "매출액",
        "sj_div": ["IS", "CIS"],
        "account_ids": ["ifrs-full_Revenue"],
        "account_names": ["Revenue", "수익(매출액)", "매출액", "영업수익"],
    },
    "operating_profit": {
        "label": "영업이익",
        "sj_div": ["IS", "CIS"],
        "account_ids": ["dart_OperatingIncomeLoss", "ifrs-full_ProfitLossFromOperatingActivities"],
        "account_names": ["Operating profit", "Operating income", "영업이익", "영업이익(손실)", "영업손실"],
    },
    "net_income": {
        "label": "당기순이익",
        "sj_div": ["IS", "CIS"],
        "account_ids": ["ifrs-full_ProfitLoss"],
        "account_names": ["Profit", "당기순이익", "당기순손익", "당기순이익(손실)"],
    },
    "total_assets": {
        "label": "자산총계",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_Assets"],
        "account_names": ["Total assets", "자산총계"],
    },
    "total_liabilities": {
        "label": "부채총계",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_Liabilities"],
        "account_names": ["Total liabilities", "부채총계"],
    },
    "total_equity": {
        "label": "자본총계",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_Equity"],
        "account_names": ["Total equity", "자본총계"],
    },
    "cash_and_cash_equivalents": {
        "label": "현금및현금성자산",
        "sj_div": ["BS"],
        "account_ids": [
            "ifrs-full_CashAndCashEquivalents",
            "ifrs-full_CashAndCashEquivalentsAtCarryingValue",
        ],
        "account_names": ["Cash and cash equivalents", "현금및현금성자산"],
    },
    "current_assets": {
        "label": "유동자산",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_CurrentAssets"],
        "account_names": ["Current assets", "유동자산"],
    },
    "current_liabilities": {
        "label": "유동부채",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_CurrentLiabilities"],
        "account_names": ["Current liabilities", "유동부채"],
    },
    "depreciation_amortization": {
        "label": "감가상각비및상각비",
        "sj_div": ["IS", "CIS", "CF"],
        "account_ids": [
            "dart_DepreciationCost",
            "ifrs-full_Depreciation",
            "ifrs-full_DepreciationAndAmortisationExpense",
            "ifrs-full_AmortisationExpense",
        ],
        "account_names": [
            "Depreciation",
            "Depreciation and amortisation expense",
            "Depreciation and amortization expense",
            "Amortisation expense",
            "감가상각비",
            "감가상각비및상각비",
            "감가상각비와상각비",
            "유무형자산상각비",
        ],
    },
    "interest_expense": {
        "label": "이자비용",
        "sj_div": ["IS", "CIS"],
        "account_ids": ["ifrs-full_FinanceCosts", "ifrs-full_InterestExpense"],
        "account_names": [
            "Interest expense",
            "Finance costs",
            "이자비용",
            "금융원가",
            "금융비용",
            "이자비용(수익)",
        ],
    },
    "operating_cash_flow": {
        "label": "영업활동현금흐름",
        "sj_div": ["CF"],
        "account_ids": ["ifrs-full_CashFlowsFromUsedInOperatingActivities"],
        "account_names": ["Cash flows from (used in) operating activities", "영업활동현금흐름"],
    },
    "investing_cash_flow": {
        "label": "투자활동현금흐름",
        "sj_div": ["CF"],
        "account_ids": ["ifrs-full_CashFlowsFromUsedInInvestingActivities"],
        "account_names": ["Cash flows from (used in) investing activities", "투자활동현금흐름"],
    },
    "financing_cash_flow": {
        "label": "재무활동현금흐름",
        "sj_div": ["CF"],
        "account_ids": ["ifrs-full_CashFlowsFromUsedInFinancingActivities"],
        "account_names": ["Cash flows from (used in) financing activities", "재무활동현금흐름"],
    },
    "trade_receivables": {
        "label": "매출채권",
        "sj_div": ["BS"],
        "account_ids": [
            "ifrs-full_CurrentTradeReceivables",
            "ifrs-full_TradeAndOtherCurrentReceivables",
            "ifrs-full_TradeReceivablesCurrent",
        ],
        "account_names": ["Trade receivables", "매출채권"],
    },
    "contract_assets": {
        "label": "계약자산",
        "sj_div": ["BS"],
        "account_ids": [
            "ifrs-full_ContractWithCustomerAssetNetCurrent",
            "ifrs-full_ContractWithCustomerAssetNet",
            "dart_ShortTermDueFromCustomersForContractWorkNet",
        ],
        "account_names": ["Contract assets", "계약자산", "단기미청구공사", "미청구공사"],
    },
    "contract_liabilities": {
        "label": "계약부채",
        "sj_div": ["BS"],
        "account_ids": [
            "ifrs-full_ContractWithCustomerLiabilityCurrent",
            "ifrs-full_ContractWithCustomerLiability",
            "dart_AdvanceReceivedsForContractWork",
        ],
        "account_names": ["Contract liabilities", "계약부채", "초과청구공사", "선수금"],
    },
    "inventories": {
        "label": "재고자산",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_Inventories"],
        "account_names": ["Inventories", "재고자산"],
    },
    "provisions": {
        "label": "충당부채",
        "sj_div": ["BS"],
        "account_ids": ["ifrs-full_Provisions", "ifrs-full_CurrentProvisions", "ifrs-full_NoncurrentProvisions"],
        "account_names": ["Provisions", "충당부채", "유동충당부채", "비유동충당부채"],
    },
    "total_borrowings": {
        "label": "차입금",
        "sj_div": ["BS"],
        "account_ids": [
            "ifrs-full_CurrentBorrowings",
            "ifrs-full_CurrentPortionOfBorrowings",
            "ifrs-full_CurrentPortionOfLongtermBorrowings",
            "ifrs-full_NoncurrentPortionOfNoncurrentLoansReceived",
            "ifrs-full_NoncurrentBorrowings",
            "ifrs-full_Borrowings",
        ],
        "account_names": ["Borrowings", "Current loans received", "Long-term borrowings", "차입금", "단기차입금", "장기차입금", "유동성장기부채"],
    },
    "capex": {
        "label": "설비투자",
        "sj_div": ["CF"],
        "account_ids": [
            "ifrs-full_PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities",
            "ifrs-full_PurchaseOfIntangibleAssetsClassifiedAsInvestingActivities",
        ],
        "account_names": [
            "Acquisition of property, plant and equipment",
            "Acquisition of intangible assets",
            "유형자산의 취득",
            "무형자산의 취득",
        ],
    },
}


RATIO_SPECS = {
    "debt_ratio": {
        "label": "부채비율",
        "numerator": "total_liabilities",
        "denominator": "total_equity",
        "format": "ratio",
    },
    "current_ratio": {
        "label": "유동비율",
        "numerator": "current_assets",
        "denominator": "current_liabilities",
        "format": "ratio",
    },
    "operating_margin": {
        "label": "영업이익률",
        "numerator": "operating_profit",
        "denominator": "revenue",
        "format": "ratio",
    },
    "net_margin": {
        "label": "순이익률",
        "numerator": "net_income",
        "denominator": "revenue",
        "format": "ratio",
    },
    "roa": {
        "label": "ROA",
        "numerator": "net_income",
        "denominator": "average_total_assets",
        "format": "ratio",
    },
    "roe": {
        "label": "ROE",
        "numerator": "net_income",
        "denominator": "average_total_equity",
        "format": "ratio",
    },
    "net_debt_to_ebitda": {
        "label": "순차입금/EBITDA",
        "numerator": "net_debt",
        "denominator": "ebitda_proxy",
        "format": "multiple",
    },
    "interest_coverage": {
        "label": "이자보상배율",
        "numerator": "operating_profit",
        "denominator": "interest_expense_abs",
        "format": "multiple",
    },
    "ebitda_interest_coverage": {
        "label": "EBITDA 이자보상배율",
        "numerator": "ebitda_proxy",
        "denominator": "interest_expense_abs",
        "format": "multiple",
    },
    "cfo_to_total_debt": {
        "label": "영업현금흐름/총차입금",
        "numerator": "operating_cash_flow",
        "denominator": "total_borrowings",
        "format": "ratio",
    },
    "contract_assets_to_revenue": {
        "label": "계약자산/매출액",
        "numerator": "contract_assets",
        "denominator": "revenue",
        "format": "ratio",
    },
    "trade_receivables_to_revenue": {
        "label": "매출채권/매출액",
        "numerator": "trade_receivables",
        "denominator": "revenue",
        "format": "ratio",
    },
    "working_capital_to_revenue": {
        "label": "운전자본/매출액",
        "numerator": "working_capital",
        "denominator": "revenue",
        "format": "ratio",
    },
}


SUMMARY_METRICS = [
    "revenue",
    "operating_profit",
    "ebitda_proxy",
    "net_income",
    "roa",
    "roe",
    "total_liabilities",
    "total_equity",
    "total_borrowings",
    "net_debt",
    "debt_ratio",
    "current_ratio",
    "operating_margin",
    "net_margin",
    "net_debt_to_ebitda",
    "interest_expense",
    "interest_coverage",
    "ebitda_interest_coverage",
    "operating_cash_flow",
    "investing_cash_flow",
    "financing_cash_flow",
    "free_cash_flow",
    "cash_and_cash_equivalents",
    "trade_receivables",
    "contract_assets",
    "contract_liabilities",
    "inventories",
    "provisions",
    "working_capital",
]


DERIVED_METRIC_LABELS = {
    "ebitda_proxy": "EBITDA proxy",
    "net_debt": "순차입금",
    "free_cash_flow": "잉여현금흐름",
    "cfo_minus_net_income": "CFO-순이익",
    "working_capital": "운전자본",
    "interest_expense_abs": "절대값 이자비용",
    "average_total_assets": "평균총자산 대용값",
    "average_total_equity": "평균자본총계 대용값",
}


METRIC_FORMATS = {
    "revenue": "amount",
    "operating_profit": "amount",
    "ebitda_proxy": "amount",
    "net_income": "amount",
    "roa": "ratio",
    "roe": "ratio",
    "total_liabilities": "amount",
    "total_equity": "amount",
    "total_borrowings": "amount",
    "net_debt": "amount",
    "debt_ratio": "ratio",
    "current_ratio": "ratio",
    "operating_margin": "ratio",
    "net_margin": "ratio",
    "net_debt_to_ebitda": "multiple",
    "interest_expense": "amount",
    "interest_coverage": "multiple",
    "ebitda_interest_coverage": "multiple",
    "operating_cash_flow": "amount",
    "investing_cash_flow": "amount",
    "financing_cash_flow": "amount",
    "free_cash_flow": "amount",
    "cash_and_cash_equivalents": "amount",
    "trade_receivables": "amount",
    "contract_assets": "amount",
    "contract_liabilities": "amount",
    "inventories": "amount",
    "provisions": "amount",
    "working_capital": "amount",
    "contract_assets_to_revenue": "ratio",
    "trade_receivables_to_revenue": "ratio",
    "working_capital_to_revenue": "ratio",
    "cfo_to_total_debt": "ratio",
    "cfo_minus_net_income": "amount",
}


@dataclass
class MetricMatch:
    value: float | None
    account_nm: str | None
    account_id: str | None
    sj_div: str | None
    candidates: list[dict]


@dataclass
class CompanyYearResult:
    input_company: str
    company: str
    corp_code: str
    stock_code: str
    year: int
    fs_div: str
    fs_label: str
    metrics: dict
    matched_accounts: dict
    notes: dict
    raw_rows: list[dict]
    red_flags: list[str]


def build_url(url: str, params: dict | None = None) -> str:
    if not params:
        return url
    return f"{url}?{urllib.parse.urlencode(params)}"


def req_get_bytes(url: str, params: dict | None = None, timeout: int = 30) -> bytes:
    request_url = build_url(url, params)
    with urllib.request.urlopen(request_url, timeout=timeout) as response:
        return response.read()


def req_get_json(url: str, params: dict | None = None, timeout: int = 30) -> dict:
    payload = req_get_bytes(url, params=params, timeout=timeout)
    return json.loads(payload.decode("utf-8"))


def clean_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_div(numerator, denominator):
    numerator = clean_number(numerator)
    denominator = clean_number(denominator)
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator


def normalize_name(name: str) -> str:
    normalized = (name or "").upper().strip()
    normalized = normalized.replace("주식회사", "")
    normalized = normalized.replace("(주)", "")
    normalized = normalized.replace("㈜", "")
    normalized = normalized.replace("주)", "")
    return re.sub(r"[\s\.\,\(\)\[\]\-_/]", "", normalized)


def normalize_account_name(name: str) -> str:
    return re.sub(r"[\s\.\,\(\)\[\]\-_/]", "", (name or "").upper())


def parse_company_names(raw_names: Iterable[str]) -> list[str]:
    names: list[str] = []
    for raw in raw_names:
        parts = raw.replace("\r", "\n").replace(",", "\n").split("\n")
        names.extend(part.strip() for part in parts if part.strip())
    return names


def year_range(start_year: int, end_year: int) -> list[int]:
    if start_year > end_year:
        raise ValueError("시작 연도는 종료 연도보다 클 수 없습니다.")
    return list(range(start_year, end_year + 1))


def download_corp_codes(api_key: str) -> list[dict]:
    response_body = req_get_bytes(CORPCODE_URL, params={"crtfc_key": api_key})
    with zipfile.ZipFile(io.BytesIO(response_body)) as zf:
        xml_name = zf.namelist()[0]
        root = ET.fromstring(zf.read(xml_name))

    rows = []
    for item in root.findall("list"):
        row = {
            "corp_code": (item.findtext("corp_code") or "").strip(),
            "corp_name": (item.findtext("corp_name") or "").strip(),
            "stock_code": (item.findtext("stock_code") or "").strip(),
            "modify_date": (item.findtext("modify_date") or "").strip(),
        }
        row["listed"] = len(row["stock_code"]) == 6
        row["norm_name"] = normalize_name(row["corp_name"])
        rows.append(row)
    return rows


def search_company_candidates(api_key: str, company_name: str, limit: int = 20) -> list[dict]:
    codes = download_corp_codes(api_key)
    norm = normalize_name(company_name)
    if not norm:
        return []

    exact = [row for row in codes if row["norm_name"] == norm]
    contains = [row for row in codes if norm in row["norm_name"] and row not in exact]
    candidates = exact + contains
    candidates.sort(key=lambda row: (row["listed"], row["modify_date"]), reverse=True)
    return [
        {
            "input_company": company_name,
            "company": row["corp_name"],
            "corp_code": row["corp_code"],
            "stock_code": row["stock_code"],
            "modify_date": row["modify_date"],
            "listed": row["listed"],
        }
        for row in candidates[:limit]
    ]


def resolve_companies(api_key: str, company_names: list[str]) -> list[dict]:
    codes = download_corp_codes(api_key)
    resolved = []
    for company in company_names:
        norm = normalize_name(company)
        matches = [row for row in codes if row["norm_name"] == norm and row["listed"]]
        if not matches:
            matches = [row for row in codes if norm in row["norm_name"] and row["listed"]]
        if not matches:
            matches = [row for row in codes if norm in row["norm_name"]]
        if not matches:
            raise ValueError(f"회사코드를 찾지 못했습니다: {company}")
        matches.sort(key=lambda row: (row["listed"], row["modify_date"]), reverse=True)
        row = matches[0]
        resolved.append(
            {
                "input_company": company,
                "company": row["corp_name"],
                "corp_code": row["corp_code"],
                "stock_code": row["stock_code"],
                "modify_date": row["modify_date"],
            }
        )
    return resolved


def fetch_financials(api_key: str, corp_code: str, year: int, fs_div: str) -> tuple[list[dict], dict]:
    payload = req_get_json(
        FS_URL,
        params={
            "crtfc_key": api_key,
            "corp_code": corp_code,
            "bsns_year": str(year),
            "reprt_code": REPORT_CODE,
            "fs_div": fs_div,
        },
    )
    status = payload.get("status")
    if status != "000":
        raise RuntimeError(f"DART error {status}: {payload.get('message')}")

    rows = payload.get("list", []) or []
    cleaned_rows = []
    for row in rows:
        copied = dict(row)
        for col, value in list(copied.items()):
            if col.endswith("_amount"):
                cleaned = str(value).replace(",", "")
                copied[col] = clean_number(cleaned) if cleaned not in {"", "nan", "None"} else None
        cleaned_rows.append(copied)
    return cleaned_rows, payload


def fetch_financials_with_fallback(api_key: str, corp_code: str, year: int) -> tuple[list[dict], str, str]:
    errors = []
    for fs_div, fs_label in ((PRIMARY_FS_DIV, "연결"), (FALLBACK_FS_DIV, "별도")):
        try:
            rows, _payload = fetch_financials(api_key, corp_code, year, fs_div)
        except Exception as exc:
            errors.append(f"{fs_label}: {exc}")
            continue
        if rows:
            return rows, fs_div, fs_label
        errors.append(f"{fs_label}: 데이터 없음")
    raise RuntimeError(" / ".join(errors) or "재무제표 데이터가 없습니다.")


def sort_candidate_rows(rows: list[dict], account_ids: list[str], account_names: list[str]) -> list[dict]:
    id_priority = {account_id: index for index, account_id in enumerate(account_ids)}
    normalized_targets = [normalize_account_name(name) for name in account_names]
    name_priority = {name: index for index, name in enumerate(normalized_targets)}

    def score(row: dict):
        account_id = str(row.get("account_id") or "")
        normalized_name = normalize_account_name(str(row.get("account_nm") or ""))
        has_exact_id = 0 if account_id in id_priority else 1
        id_rank = id_priority.get(account_id, 999)
        has_exact_name = 0 if normalized_name in name_priority else 1
        name_rank = name_priority.get(normalized_name, 999)
        contains_rank = 999
        for index, target in enumerate(normalized_targets):
            if target and target in normalized_name:
                contains_rank = index
                break
        account_detail = str(row.get("account_detail") or "")
        detail_penalty = 1 if account_detail and account_detail != "-" else 0
        ord_value = int(str(row.get("ord") or "9999"))
        return (has_exact_id, id_rank, has_exact_name, name_rank, contains_rank, detail_penalty, ord_value)

    return sorted(rows, key=score)


def find_amount(rows: list[dict], sj_div: str | list[str], account_ids: list[str], account_names: list[str]) -> MetricMatch:
    sj_divs = sj_div if isinstance(sj_div, list) else [sj_div]
    scoped_rows = [row for row in rows if row.get("sj_div") in sj_divs]
    if not scoped_rows:
        return MetricMatch(None, None, None, None, [])

    candidates = []
    normalized_targets = [normalize_account_name(name) for name in account_names]
    for row in scoped_rows:
        account_id = str(row.get("account_id") or "")
        normalized_name = normalize_account_name(str(row.get("account_nm") or ""))
        matched = account_id in account_ids or normalized_name in normalized_targets
        if not matched:
            matched = any(target and target in normalized_name for target in normalized_targets)
        if matched:
            candidates.append(row)

    if not candidates:
        return MetricMatch(None, None, None, None, [])

    sorted_candidates = sort_candidate_rows(candidates, account_ids, account_names)
    best = sorted_candidates[0]
    alternative_rows = [
        {
            "account_id": row.get("account_id"),
            "account_nm": row.get("account_nm"),
            "sj_div": row.get("sj_div"),
            "amount": clean_number(row.get("thstrm_amount")),
        }
        for row in sorted_candidates[1:4]
    ]
    return MetricMatch(
        clean_number(best.get("thstrm_amount")),
        best.get("account_nm"),
        best.get("account_id"),
        best.get("sj_div"),
        alternative_rows,
    )


def source_entry(match: MetricMatch | None, label: str) -> dict:
    match = match or MetricMatch(None, None, None, None, [])
    return {
        "label": match.account_nm or label,
        "account_id": match.account_id,
        "sj_div": match.sj_div,
        "value": match.value,
        "alternatives": match.candidates,
    }


def explain_ratio(metric_key: str, numerator_value, denominator_value) -> str:
    label = RATIO_SPECS[metric_key]["label"]
    if numerator_value is None:
        return f"분자 값이 없어 {label}을 계산하지 못했습니다."
    if denominator_value is None:
        return f"분모 값이 없어 {label}을 계산하지 못했습니다."
    if clean_number(denominator_value) == 0:
        return f"분모가 0이라 {label}을 계산하지 못했습니다."
    return "정상 계산되었습니다."


def value_or_blank_reason(value, success_message: str, missing_message: str) -> str:
    return success_message if value is not None else missing_message


def build_red_flags(metric_values: dict, notes: dict) -> list[str]:
    flags = []
    if metric_values.get("debt_ratio") is not None and metric_values["debt_ratio"] >= 2:
        flags.append("예비 적색신호: 부채비율이 200% 이상입니다. 인적 검토가 필요합니다.")
    if metric_values.get("interest_coverage") is not None and metric_values["interest_coverage"] < 1:
        flags.append("예비 적색신호: 이자보상배율이 1배 미만입니다. 인적 검토가 필요합니다.")
    if metric_values.get("operating_cash_flow") is not None and metric_values["operating_cash_flow"] < 0:
        flags.append("예비 적색신호: 영업현금흐름이 음수입니다. 인적 검토가 필요합니다.")
    if metric_values.get("current_ratio") is not None and metric_values["current_ratio"] < 1:
        flags.append("예비 적색신호: 유동비율이 100% 미만입니다. 인적 검토가 필요합니다.")
    if metric_values.get("working_capital") is not None and metric_values["working_capital"] < 0:
        flags.append("예비 적색신호: 운전자본이 음수입니다. 인적 검토가 필요합니다.")
    if metric_values.get("net_debt_to_ebitda") is not None and metric_values["net_debt_to_ebitda"] > 5:
        flags.append("예비 적색신호: 순차입금/EBITDA가 5배를 초과합니다. 인적 검토가 필요합니다.")
    if not flags:
        flags.append("자동 적색신호 없음. 인적 검토는 여전히 필요합니다.")
    for metric_key in ("ebitda_proxy", "total_borrowings", "contract_assets", "contract_liabilities", "provisions"):
        note = notes.get(metric_key, "")
        if "찾지 못했습니다" in note:
            flags.append(f"참고: {note}")
    return flags


def normalize_statement(
    statement_rows: list[dict],
    company_row: dict,
    year: int,
    fs_div: str,
    fs_label: str,
) -> tuple[CompanyYearResult, list[dict]]:
    metric_values = {}
    matched_accounts = {}
    notes = {}
    matched_rows = []

    for metric_key, spec in METRIC_SPECS.items():
        match = find_amount(statement_rows, spec["sj_div"], spec["account_ids"], spec["account_names"])
        metric_values[metric_key] = match.value
        matched_accounts[metric_key] = source_entry(match, spec["label"])
        if match.value is None:
            notes[metric_key] = f"{spec['label']} 계정을 찾지 못했습니다."
        else:
            notes[metric_key] = "정상 조회되었습니다."
        matched_rows.append(
            {
                "입력회사명": company_row["input_company"],
                "회사명": company_row["company"],
                "사업연도": year,
                "재무제표구분": fs_label,
                "지표키": metric_key,
                "지표명": spec["label"],
                "사용계정명": match.account_nm,
                "사용계정ID": match.account_id,
                "표구분": match.sj_div,
                "사용금액": match.value,
                "대안후보": " | ".join(
                    f"{candidate.get('account_nm')} ({candidate.get('account_id')})"
                    for candidate in match.candidates
                ),
            }
        )

    total_borrowings_match = matched_accounts.get("total_borrowings")
    operating_profit = metric_values.get("operating_profit")
    depreciation_amortization = metric_values.get("depreciation_amortization")
    cash = metric_values.get("cash_and_cash_equivalents")
    total_borrowings = metric_values.get("total_borrowings")
    operating_cash_flow = metric_values.get("operating_cash_flow")
    net_income = metric_values.get("net_income")
    capex = metric_values.get("capex")
    total_assets = metric_values.get("total_assets")
    total_equity = metric_values.get("total_equity")
    interest_expense = metric_values.get("interest_expense")
    current_assets = metric_values.get("current_assets")
    current_liabilities = metric_values.get("current_liabilities")
    contract_assets = metric_values.get("contract_assets")
    trade_receivables = metric_values.get("trade_receivables")
    revenue = metric_values.get("revenue")

    metric_values["ebitda_proxy"] = None if operating_profit is None or depreciation_amortization is None else operating_profit + depreciation_amortization
    notes["ebitda_proxy"] = (
        "정상 계산되었습니다."
        if metric_values["ebitda_proxy"] is not None
        else "영업이익 또는 감가상각비및상각비를 찾지 못해 EBITDA proxy를 계산하지 못했습니다."
    )
    matched_accounts["ebitda_proxy"] = {
        "numerator": matched_accounts.get("operating_profit"),
        "denominator": matched_accounts.get("depreciation_amortization"),
    }
    matched_rows.append(
        {
            "입력회사명": company_row["input_company"],
            "회사명": company_row["company"],
            "사업연도": year,
            "재무제표구분": fs_label,
            "지표키": "ebitda_proxy",
            "지표명": "EBITDA proxy",
            "사용계정명": f"{METRIC_SPECS['operating_profit']['label']} + {METRIC_SPECS['depreciation_amortization']['label']}",
            "사용계정ID": f"{matched_accounts['operating_profit'].get('account_id')} + {matched_accounts['depreciation_amortization'].get('account_id')}",
            "표구분": f"{matched_accounts['operating_profit'].get('sj_div')} + {matched_accounts['depreciation_amortization'].get('sj_div')}",
            "사용금액": metric_values["ebitda_proxy"],
            "대안후보": "",
        }
    )

    metric_values["net_debt"] = None if total_borrowings is None or cash is None else total_borrowings - cash
    notes["net_debt"] = (
        "정상 계산되었습니다."
        if metric_values["net_debt"] is not None
        else "차입금 또는 현금및현금성자산을 찾지 못해 순차입금을 계산하지 못했습니다."
    )
    matched_accounts["net_debt"] = {
        "numerator": matched_accounts.get("total_borrowings"),
        "denominator": matched_accounts.get("cash_and_cash_equivalents"),
    }
    matched_rows.append(
        {
            "입력회사명": company_row["input_company"],
            "회사명": company_row["company"],
            "사업연도": year,
            "재무제표구분": fs_label,
            "지표키": "net_debt",
            "지표명": "순차입금",
            "사용계정명": f"{METRIC_SPECS['total_borrowings']['label']} - {METRIC_SPECS['cash_and_cash_equivalents']['label']}",
            "사용계정ID": f"{matched_accounts['total_borrowings'].get('account_id')} - {matched_accounts['cash_and_cash_equivalents'].get('account_id')}",
            "표구분": f"{matched_accounts['total_borrowings'].get('sj_div')} - {matched_accounts['cash_and_cash_equivalents'].get('sj_div')}",
            "사용금액": metric_values["net_debt"],
            "대안후보": "",
        }
    )

    metric_values["interest_expense_abs"] = abs(interest_expense) if interest_expense is not None else None
    matched_accounts["interest_expense_abs"] = matched_accounts.get("interest_expense")
    notes["interest_expense_abs"] = notes.get("interest_expense", "")

    metric_values["free_cash_flow"] = None if operating_cash_flow is None or capex is None else operating_cash_flow - abs(capex)
    notes["free_cash_flow"] = (
        "정상 계산되었습니다."
        if metric_values["free_cash_flow"] is not None
        else "영업현금흐름 또는 설비투자를 찾지 못해 잉여현금흐름을 계산하지 못했습니다."
    )
    matched_accounts["free_cash_flow"] = {
        "numerator": matched_accounts.get("operating_cash_flow"),
        "denominator": matched_accounts.get("capex"),
    }
    matched_rows.append(
        {
            "입력회사명": company_row["input_company"],
            "회사명": company_row["company"],
            "사업연도": year,
            "재무제표구분": fs_label,
            "지표키": "free_cash_flow",
            "지표명": "잉여현금흐름",
            "사용계정명": f"{METRIC_SPECS['operating_cash_flow']['label']} - {METRIC_SPECS['capex']['label']}",
            "사용계정ID": f"{matched_accounts['operating_cash_flow'].get('account_id')} - {matched_accounts['capex'].get('account_id')}",
            "표구분": f"{matched_accounts['operating_cash_flow'].get('sj_div')} - {matched_accounts['capex'].get('sj_div')}",
            "사용금액": metric_values["free_cash_flow"],
            "대안후보": "",
        }
    )

    metric_values["cfo_minus_net_income"] = None if operating_cash_flow is None or net_income is None else operating_cash_flow - net_income
    notes["cfo_minus_net_income"] = (
        "정상 계산되었습니다."
        if metric_values["cfo_minus_net_income"] is not None
        else "영업현금흐름 또는 당기순이익을 찾지 못해 CFO - 순이익을 계산하지 못했습니다."
    )
    matched_accounts["cfo_minus_net_income"] = {
        "numerator": matched_accounts.get("operating_cash_flow"),
        "denominator": matched_accounts.get("net_income"),
    }
    matched_rows.append(
        {
            "입력회사명": company_row["input_company"],
            "회사명": company_row["company"],
            "사업연도": year,
            "재무제표구분": fs_label,
            "지표키": "cfo_minus_net_income",
            "지표명": "CFO-순이익",
            "사용계정명": f"{METRIC_SPECS['operating_cash_flow']['label']} - {METRIC_SPECS['net_income']['label']}",
            "사용계정ID": f"{matched_accounts['operating_cash_flow'].get('account_id')} - {matched_accounts['net_income'].get('account_id')}",
            "표구분": f"{matched_accounts['operating_cash_flow'].get('sj_div')} - {matched_accounts['net_income'].get('sj_div')}",
            "사용금액": metric_values["cfo_minus_net_income"],
            "대안후보": "",
        }
    )

    metric_values["working_capital"] = None if current_assets is None or current_liabilities is None else current_assets - current_liabilities
    notes["working_capital"] = (
        "정상 계산되었습니다."
        if metric_values["working_capital"] is not None
        else "유동자산 또는 유동부채를 찾지 못해 운전자본을 계산하지 못했습니다."
    )
    matched_accounts["working_capital"] = {
        "numerator": matched_accounts.get("current_assets"),
        "denominator": matched_accounts.get("current_liabilities"),
    }
    matched_rows.append(
        {
            "입력회사명": company_row["input_company"],
            "회사명": company_row["company"],
            "사업연도": year,
            "재무제표구분": fs_label,
            "지표키": "working_capital",
            "지표명": "운전자본",
            "사용계정명": f"{METRIC_SPECS['current_assets']['label']} - {METRIC_SPECS['current_liabilities']['label']}",
            "사용계정ID": f"{matched_accounts['current_assets'].get('account_id')} - {matched_accounts['current_liabilities'].get('account_id')}",
            "표구분": f"{matched_accounts['current_assets'].get('sj_div')} - {matched_accounts['current_liabilities'].get('sj_div')}",
            "사용금액": metric_values["working_capital"],
            "대안후보": "",
        }
    )

    metric_values["average_total_assets"] = total_assets
    notes["average_total_assets"] = "전기 평균 자산이 없어 당기말 자산총계를 대용값으로 사용했습니다." if total_assets is not None else "자산총계를 찾지 못했습니다."
    matched_accounts["average_total_assets"] = matched_accounts.get("total_assets")

    metric_values["average_total_equity"] = total_equity
    notes["average_total_equity"] = "전기 평균 자본이 없어 당기말 자본총계를 대용값으로 사용했습니다." if total_equity is not None else "자본총계를 찾지 못했습니다."
    matched_accounts["average_total_equity"] = matched_accounts.get("total_equity")

    for metric_key, ratio_spec in RATIO_SPECS.items():
        numerator_key = ratio_spec["numerator"]
        denominator_key = ratio_spec["denominator"]
        numerator_value = metric_values.get(numerator_key)
        denominator_value = metric_values.get(denominator_key)
        metric_values[metric_key] = safe_div(numerator_value, denominator_value)
        matched_accounts[metric_key] = {
            "numerator": matched_accounts.get(numerator_key),
            "denominator": matched_accounts.get(denominator_key),
        }
        notes[metric_key] = explain_ratio(metric_key, numerator_value, denominator_value)
        matched_rows.append(
            {
                "입력회사명": company_row["input_company"],
                "회사명": company_row["company"],
                "사업연도": year,
                "재무제표구분": fs_label,
                "지표키": metric_key,
                "지표명": ratio_spec["label"],
                "사용계정명": f"{metric_label(numerator_key)} / {metric_label(denominator_key)}",
                "사용계정ID": f"{matched_accounts[numerator_key].get('account_id')} / {matched_accounts[denominator_key].get('account_id')}",
                "표구분": f"{matched_accounts[numerator_key].get('sj_div')} / {matched_accounts[denominator_key].get('sj_div')}",
                "사용금액": metric_values[metric_key],
                "대안후보": "",
            }
        )

    red_flags = build_red_flags(metric_values, notes)
    result = CompanyYearResult(
        input_company=company_row["input_company"],
        company=company_row["company"],
        corp_code=company_row["corp_code"],
        stock_code=company_row["stock_code"],
        year=year,
        fs_div=fs_div,
        fs_label=fs_label,
        metrics=metric_values,
        matched_accounts=matched_accounts,
        notes=notes,
        raw_rows=statement_rows,
        red_flags=red_flags,
    )
    return result, matched_rows


def build_raw_fs_rows(company_row: dict, year: int, fs_label: str, rows: list[dict]) -> list[dict]:
    out = []
    for row in rows:
        out.append(
            {
                "입력회사명": company_row["input_company"],
                "회사명": company_row["company"],
                "corp_code": company_row["corp_code"],
                "stock_code": company_row["stock_code"],
                "사업연도": year,
                "재무제표구분": fs_label,
                "statement_type": row.get("sj_div"),
                "statement_name": row.get("sj_nm"),
                "account_id": row.get("account_id"),
                "account_nm": row.get("account_nm"),
                "account_detail": row.get("account_detail"),
                "amount": row.get("thstrm_amount"),
                "currency": row.get("currency"),
                "ord": row.get("ord"),
                "rcept_no": row.get("rcept_no"),
            }
        )
    return out


def evaluate_company_year(api_key: str, company_row: dict, year: int) -> tuple[CompanyYearResult, list[dict], list[dict]]:
    rows, fs_div, fs_label = fetch_financials_with_fallback(api_key, company_row["corp_code"], year)
    result, matched_rows = normalize_statement(rows, company_row, year, fs_div, fs_label)
    raw_rows = build_raw_fs_rows(company_row, year, fs_label, rows)
    return result, matched_rows, raw_rows


def run_multi_company_analysis(api_key: str, companies: list[dict], years: list[int]) -> dict:
    results: list[CompanyYearResult] = []
    matched_rows: list[dict] = []
    raw_rows: list[dict] = []
    errors: list[dict] = []

    for company in companies:
        if company.get("unresolved"):
            for year in years:
                errors.append(
                    {
                        "입력회사명": company["input_company"],
                        "회사명": company["company"],
                        "corp_code": company.get("corp_code", ""),
                        "stock_code": company.get("stock_code", ""),
                        "사업연도": year,
                        "오류내용": "회사 검색 결과가 없어 재무제표를 조회하지 못했습니다.",
                    }
                )
            continue
        for year in years:
            try:
                result, matched, raw = evaluate_company_year(api_key, company, year)
                results.append(result)
                matched_rows.extend(matched)
                raw_rows.extend(raw)
            except Exception as exc:
                errors.append(
                    {
                        "입력회사명": company["input_company"],
                        "회사명": company["company"],
                        "corp_code": company["corp_code"],
                        "stock_code": company["stock_code"],
                        "사업연도": year,
                        "오류내용": str(exc),
                    }
                )
    return {
        "results": results,
        "matched_rows": matched_rows,
        "raw_rows": raw_rows,
        "errors": errors,
    }


def format_number(value, kind: str):
    if value is None:
        return None
    if kind == "ratio":
        return value
    return value


def build_summary_rows(results: list[CompanyYearResult]) -> list[dict]:
    latest_by_company: dict[str, CompanyYearResult] = {}
    for result in results:
        current = latest_by_company.get(result.company)
        if current is None or result.year > current.year:
            latest_by_company[result.company] = result

    rows = []
    for company_name, result in sorted(latest_by_company.items()):
        red_flags = result.red_flags or []
        severe_count = sum(1 for flag in red_flags if "200% 이상" in flag or "1배 미만" in flag or "음수" in flag or "5배를 초과" in flag)
        missing_notes = sorted({note for note in result.notes.values() if "찾지 못했습니다" in note or "계산하지 못했습니다" in note})
        row = {
            "회사명": result.company,
            "입력회사명": result.input_company,
            "종목코드": result.stock_code,
            "corp_code": result.corp_code,
            "기준연도": result.year,
            "재무제표구분": result.fs_label,
            "연결대체여부": "별도 사용" if result.fs_div == FALLBACK_FS_DIV else "연결 사용",
            "예비적색신호수": len(red_flags),
            "중요적색신호수": severe_count,
            "누락데이터메모": " | ".join(missing_notes[:6]),
            "예비적색신호": " | ".join(red_flags),
        }
        for metric_key in SUMMARY_METRICS:
            label = metric_label(metric_key)
            row[label] = result.metrics.get(metric_key)
        rows.append(row)
    return rows


def build_financial_metric_rows(results: list[CompanyYearResult]) -> list[dict]:
    ordered_metric_keys = [
        "revenue",
        "operating_profit",
        "ebitda_proxy",
        "operating_margin",
        "net_income",
        "net_margin",
        "roa",
        "roe",
        "total_assets",
        "total_liabilities",
        "total_equity",
        "debt_ratio",
        "total_borrowings",
        "net_debt",
        "net_debt_to_ebitda",
        "interest_expense",
        "interest_coverage",
        "ebitda_interest_coverage",
        "current_assets",
        "current_liabilities",
        "current_ratio",
        "cash_and_cash_equivalents",
        "operating_cash_flow",
        "investing_cash_flow",
        "financing_cash_flow",
        "free_cash_flow",
        "cfo_to_total_debt",
        "cfo_minus_net_income",
        "trade_receivables",
        "contract_assets",
        "contract_liabilities",
        "inventories",
        "provisions",
        "contract_assets_to_revenue",
        "trade_receivables_to_revenue",
        "working_capital",
        "working_capital_to_revenue",
    ]
    rows = []
    for result in results:
        row = {
            "회사명": result.company,
            "입력회사명": result.input_company,
            "종목코드": result.stock_code,
            "corp_code": result.corp_code,
            "사업연도": result.year,
            "재무제표구분": result.fs_label,
            "예비적색신호수": len(result.red_flags or []),
            "중요적색신호수": sum(1 for flag in (result.red_flags or []) if "200% 이상" in flag or "1배 미만" in flag or "음수" in flag or "5배를 초과" in flag),
        }
        for metric_key in ordered_metric_keys:
            row[metric_label(metric_key)] = result.metrics.get(metric_key)
        row["누락/주의메모"] = " | ".join(sorted({note for note in result.notes.values() if note and note != "정상 조회되었습니다." and note != "정상 계산되었습니다."})[:8])
        rows.append(row)
    rows.sort(key=lambda item: (item["회사명"], item["사업연도"]))
    return rows


def build_metric_definition_rows() -> list[dict]:
    return [
        {"지표명": "Revenue", "한글명": "매출액", "정의": "OpenDART 재무제표에서 매출 또는 수익(매출액) 계정을 사용한 재무 스크리닝 지표", "비고": "인적 검토 필요"},
        {"지표명": "Operating profit", "한글명": "영업이익", "정의": "영업이익(손실) 계정", "비고": "인적 검토 필요"},
        {"지표명": "EBITDA proxy", "한글명": "EBITDA proxy", "정의": "영업이익 + 감가상각비및상각비", "비고": "감가상각 계정 매칭 민감"},
        {"지표명": "Operating margin", "한글명": "영업이익률", "정의": "영업이익 / 매출액", "비고": "인적 검토 필요"},
        {"지표명": "Net income", "한글명": "당기순이익", "정의": "당기순이익(손실) 계정", "비고": "인적 검토 필요"},
        {"지표명": "Net margin", "한글명": "순이익률", "정의": "당기순이익 / 매출액", "비고": "인적 검토 필요"},
        {"지표명": "ROA", "한글명": "ROA", "정의": "당기순이익 / 평균총자산. 현재는 평균값 대신 당기말 자산총계 대용", "비고": "보수적 해석 필요"},
        {"지표명": "ROE", "한글명": "ROE", "정의": "당기순이익 / 평균자본총계. 현재는 평균값 대신 당기말 자본총계 대용", "비고": "보수적 해석 필요"},
        {"지표명": "Debt ratio", "한글명": "부채비율", "정의": "부채총계 / 자본총계", "비고": "인적 검토 필요"},
        {"지표명": "Total borrowings", "한글명": "총차입금", "정의": "단기차입금, 장기차입금, 유동성장기부채 등 이자발생성 차입 계정 우선 합산 개념", "비고": "계정 매칭 민감"},
        {"지표명": "Net debt", "한글명": "순차입금", "정의": "총차입금 - 현금및현금성자산", "비고": "인적 검토 필요"},
        {"지표명": "Net debt / EBITDA", "한글명": "순차입금/EBITDA", "정의": "순차입금 / EBITDA proxy", "비고": "감가상각 계정 없으면 공란"},
        {"지표명": "Interest coverage", "한글명": "이자보상배율", "정의": "영업이익 / 절대값 이자비용", "비고": "금융원가를 이자비용 proxy로 사용 가능"},
        {"지표명": "EBITDA interest coverage", "한글명": "EBITDA 이자보상배율", "정의": "EBITDA proxy / 절대값 이자비용", "비고": "감가상각 계정 민감"},
        {"지표명": "CFO / total debt", "한글명": "영업현금흐름/총차입금", "정의": "영업활동현금흐름 / 총차입금", "비고": "인적 검토 필요"},
        {"지표명": "Free cash flow", "한글명": "잉여현금흐름", "정의": "영업활동현금흐름 - 설비투자", "비고": "capex 계정 민감"},
        {"지표명": "CFO - net income", "한글명": "CFO-순이익", "정의": "영업활동현금흐름 - 당기순이익", "비고": "인적 검토 필요"},
        {"지표명": "Contract assets / revenue", "한글명": "계약자산/매출액", "정의": "계약자산 / 매출액", "비고": "건설사에 유용, 계정 매칭 민감"},
        {"지표명": "Trade receivables / revenue", "한글명": "매출채권/매출액", "정의": "매출채권 / 매출액", "비고": "인적 검토 필요"},
        {"지표명": "Working capital", "한글명": "운전자본", "정의": "유동자산 - 유동부채", "비고": "인적 검토 필요"},
        {"지표명": "Working capital / revenue", "한글명": "운전자본/매출액", "정의": "운전자본 / 매출액", "비고": "인적 검토 필요"},
        {"지표명": "Program positioning", "한글명": "프로그램 성격", "정의": "정식 신용등급이 아닌 preliminary financial screening 및 preliminary red flag 도구", "비고": "human review required"},
    ]


def metric_label(metric_key: str) -> str:
    return (
        METRIC_SPECS.get(metric_key, {}).get("label")
        or RATIO_SPECS.get(metric_key, {}).get("label")
        or DERIVED_METRIC_LABELS.get(metric_key)
        or metric_key
    )


def build_red_flag_rows(results: list[CompanyYearResult]) -> list[dict]:
    rows = []
    for result in sorted(results, key=lambda item: (item.company, item.year)):
        flags = result.red_flags or ["자동 적색신호 없음. 인적 검토는 여전히 필요합니다."]
        for index, flag in enumerate(flags, start=1):
            severity = "중요" if any(token in flag for token in ("200% 이상", "1배 미만", "음수", "5배를 초과")) else "일반"
            rows.append(
                {
                    "회사명": result.company,
                    "사업연도": result.year,
                    "재무제표구분": result.fs_label,
                    "적색신호순번": index,
                    "중요도": severity,
                    "적색신호내용": flag,
                }
            )
    return rows


def autosize_worksheet(ws):
    from openpyxl.utils import get_column_letter

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 40)


def apply_header_style(ws):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def append_sheet(ws, rows: list[dict], ratio_columns: set[str] | None = None):
    ratio_columns = ratio_columns or set()
    if not rows:
        ws.append(["데이터 없음"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    apply_header_style(ws)
    autosize_worksheet(ws)
    for header in headers:
        if header in ratio_columns:
            col_idx = headers.index(header) + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "0.0%"
        elif header == "값" and any("율" in str(row.get("지표명", "")) for row in rows):
            col_idx = headers.index(header) + 1
            for row_idx in range(2, ws.max_row + 1):
                metric_name = ws.cell(row_idx, headers.index("지표명") + 1).value if "지표명" in headers else ""
                if metric_name and "율" in str(metric_name):
                    ws.cell(row_idx, col_idx).number_format = "0.0%"
                else:
                    ws.cell(row_idx, col_idx).number_format = '#,##0;[Red](#,##0);-'
        elif header in {"매출액", "영업이익", "당기순이익", "자산총계", "부채총계", "자본총계"}:
            col_idx = headers.index(header) + 1
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0;[Red](#,##0);-'


def style_review_sheet(ws, metric_columns: dict[str, str] | None = None):
    from openpyxl.styles import Alignment, PatternFill

    metric_columns = metric_columns or {}
    missing_fill = PatternFill("solid", fgColor="FFF7E6")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.value is None:
                cell.fill = missing_fill
    headers = [cell.value for cell in ws[1]]
    for header, fmt in metric_columns.items():
        if header not in headers:
            continue
        col_idx = headers.index(header) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is None:
                continue
            if fmt == "amount":
                cell.number_format = '#,##0;[Red](#,##0);-'
            elif fmt == "ratio":
                cell.number_format = "0.0%"
            elif fmt == "multiple":
                cell.number_format = '0.00x'


def build_chart_data(ws, results: list[CompanyYearResult]):
    metric_map = {
        "매출액 추이": ("revenue", "amount"),
        "영업이익률 추이": ("operating_margin", "ratio"),
        "부채비율 추이": ("debt_ratio", "ratio"),
        "영업현금흐름 추이": ("operating_cash_flow", "amount"),
        "계약자산/매출액 추이": ("contract_assets_to_revenue", "ratio"),
    }
    companies = sorted({result.company for result in results})
    years = sorted({result.year for result in results})
    result_map = {(result.company, result.year): result for result in results}

    current_row = 1
    for chart_title, (metric_key, _format_kind) in metric_map.items():
        ws.cell(current_row, 1).value = chart_title
        ws.cell(current_row + 1, 1).value = "연도"
        for index, company in enumerate(companies, start=2):
            ws.cell(current_row + 1, index).value = company
        for row_offset, year in enumerate(years, start=2):
            ws.cell(current_row + row_offset, 1).value = year
            for col_offset, company in enumerate(companies, start=2):
                result = result_map.get((company, year))
                value = result.metrics.get(metric_key) if result else None
                if value is not None and _format_kind == "amount":
                    value = value / 100000000
                ws.cell(current_row + row_offset, col_offset).value = value
        current_row += len(years) + 4

    ws.cell(current_row, 1).value = "회사별 적색신호"
    ws.cell(current_row + 1, 1).value = "회사명"
    ws.cell(current_row + 1, 2).value = "적색신호수"
    ws.cell(current_row + 1, 3).value = "중요적색신호수"
    for index, company in enumerate(companies, start=current_row + 2):
        company_results = [result for result in results if result.company == company]
        total_flags = sum(len(result.red_flags or []) for result in company_results)
        severe_flags = sum(
            sum(1 for flag in (result.red_flags or []) if "200% 이상" in flag or "1배 미만" in flag or "음수" in flag or "5배를 초과" in flag)
            for result in company_results
        )
        ws.cell(index, 1).value = company
        ws.cell(index, 2).value = total_flags
        ws.cell(index, 3).value = severe_flags


def add_line_chart(target_ws, data_ws, title: str, anchor: str, block_start_row: int, company_count: int, year_count: int, y_axis_title: str):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 10
    chart.width = 16
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "사업연도"
    chart.legend.position = "r"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    data = Reference(
        data_ws,
        min_col=2,
        max_col=1 + company_count,
        min_row=block_start_row + 1,
        max_row=block_start_row + 1 + year_count,
    )
    cats = Reference(
        data_ws,
        min_col=1,
        min_row=block_start_row + 2,
        max_row=block_start_row + 1 + year_count,
    )
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    target_ws.add_chart(chart, anchor)


def add_bar_chart(target_ws, data_ws, title: str, anchor: str, category_col: int, value_col: int, header_row: int, start_row: int, end_row: int, y_axis_title: str):
    from openpyxl.chart import BarChart
    from openpyxl.chart import Reference
    from openpyxl.chart.label import DataLabelList

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.height = 10
    chart.width = 16
    chart.x_axis.title = "회사명"
    chart.y_axis.title = y_axis_title
    chart.legend = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    data = Reference(data_ws, min_col=value_col, min_row=header_row, max_row=end_row)
    cats = Reference(data_ws, min_col=category_col, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    target_ws.add_chart(chart, anchor)


def add_charts_sheet(charts_ws, chart_data_ws):
    charts_ws["A1"] = "재무 스크리닝 차트"
    from openpyxl.styles import Font
    charts_ws["A1"].font = Font(bold=True, size=14)
    charts_ws["A2"] = "각 차트는 preliminary financial screening용이며 human review required."
    charts_ws.sheet_view.showGridLines = False
    companies = []
    col = 2
    while chart_data_ws.cell(2, col).value:
        companies.append(chart_data_ws.cell(2, col).value)
        col += 1
    years = []
    row = 3
    while chart_data_ws.cell(row, 1).value:
        years.append(chart_data_ws.cell(row, 1).value)
        row += 1
    company_count = len(companies)
    year_count = len(years)

    add_line_chart(charts_ws, chart_data_ws, "매출액 추이", "A4", 1, company_count, year_count, "금액 (억원)")
    add_line_chart(charts_ws, chart_data_ws, "영업이익률 추이", "R4", 1 + year_count + 4, company_count, year_count, "비율 (%)")
    add_line_chart(charts_ws, chart_data_ws, "부채비율 추이", "A24", 1 + (year_count + 4) * 2, company_count, year_count, "비율 (%)")
    add_line_chart(charts_ws, chart_data_ws, "영업현금흐름 추이", "R24", 1 + (year_count + 4) * 3, company_count, year_count, "금액 (억원)")
    add_line_chart(charts_ws, chart_data_ws, "계약자산/매출액 추이", "A44", 1 + (year_count + 4) * 4, company_count, year_count, "비율 (%)")

    company_summary_start = 1 + (year_count + 4) * 5
    add_bar_chart(charts_ws, chart_data_ws, "회사별 적색신호 수", "R44", 1, 2, company_summary_start + 1, company_summary_start + 2, company_summary_start + 1 + company_count, "건수")
    add_bar_chart(charts_ws, chart_data_ws, "회사별 중요 적색신호 수", "A64", 1, 3, company_summary_start + 1, company_summary_start + 2, company_summary_start + 1 + company_count, "건수")


def write_workbook(analysis: dict, output_path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("엑셀 저장을 위해 openpyxl 설치가 필요합니다. 예: pip install openpyxl") from exc

    wb = Workbook()

    charts_ws = wb.active
    charts_ws.title = "Charts"

    summary_ws = wb.create_sheet("Summary")
    summary_rows = build_summary_rows(analysis["results"])
    year_values = sorted({result.year for result in analysis["results"]})
    company_names = sorted({result.company for result in analysis["results"]})
    summary_ws["A1"] = "OpenDART 예비 재무 스크리닝 요약"
    summary_ws["A1"].font = Font(bold=True, size=15)
    summary_ws["A2"] = "회사 목록"
    summary_ws["B2"] = ", ".join(company_names)
    summary_ws["A3"] = "연도 범위"
    summary_ws["B3"] = f"{min(year_values)} ~ {max(year_values)}" if year_values else "-"
    summary_ws["A4"] = "안내"
    summary_ws["B4"] = "정식 신용등급이 아닌 preliminary financial screening 결과이며 human review required."
    append_sheet(
        summary_ws,
        summary_rows,
        ratio_columns={"부채비율", "유동비율", "영업이익률", "순이익률"},
    )
    summary_ws.freeze_panes = "A6"

    metrics_ws = wb.create_sheet("Financial Metrics")
    metric_rows = build_financial_metric_rows(analysis["results"])
    append_sheet(metrics_ws, metric_rows)
    metrics_ws.freeze_panes = "A2"
    style_review_sheet(metrics_ws, {metric_label(key): fmt for key, fmt in METRIC_FORMATS.items()})

    matched_ws = wb.create_sheet("Matched Accounts")
    append_sheet(matched_ws, analysis["matched_rows"])
    matched_ws.freeze_panes = "A2"

    raw_ws = wb.create_sheet("Raw FS Data")
    append_sheet(raw_ws, analysis["raw_rows"])
    raw_ws.freeze_panes = "A2"

    errors_ws = wb.create_sheet("Errors - Missing Data")
    append_sheet(errors_ws, analysis["errors"])

    red_flags_ws = wb.create_sheet("Red Flag Details")
    append_sheet(red_flags_ws, build_red_flag_rows(analysis["results"]))

    definitions_ws = wb.create_sheet("Metric Definitions")
    append_sheet(definitions_ws, build_metric_definition_rows())

    if metric_rows:
        chart_data_ws = wb.create_sheet("Chart Data")
        build_chart_data(chart_data_ws, analysis["results"])
        add_charts_sheet(charts_ws, chart_data_ws)
        chart_data_ws.sheet_state = "hidden"

    wb.save(output_path)


def evaluate_companies(api_key: str, company_names: list[str], start_year: int, end_year: int, output_path: Path):
    companies = resolve_companies(api_key, company_names)
    analysis = run_multi_company_analysis(api_key, companies, year_range(start_year, end_year))
    write_workbook(analysis, output_path)
    return analysis


def print_results(analysis: dict, output_path: Path):
    print("\nOpenDART 비교 분석 결과")
    print("-" * 72)
    print(f"성공 건수: {len(analysis['results'])}")
    print(f"오류 건수: {len(analysis['errors'])}")
    print(f"엑셀 저장: {output_path.resolve()}")


def build_parser():
    parser = argparse.ArgumentParser(description="OpenDART 다중 회사/다중 연도 재무 비교 도구")
    parser.add_argument("companies", nargs="*", help="회사명 목록. 쉼표로 여러 회사를 입력할 수 있습니다.")
    parser.add_argument("--api-key", default=os.getenv("DART_API_KEY") or BUILT_IN_API_KEY, help=argparse.SUPPRESS)
    parser.add_argument("--start-year", type=int, required=True, help="시작 연도")
    parser.add_argument("--end-year", type=int, required=True, help="종료 연도")
    parser.add_argument("--output", default="dart_company_comparison.xlsx", help="결과 엑셀 파일명")
    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()
    company_names = parse_company_names(args.companies)
    if not company_names:
        company_names = parse_company_names([input("조회할 회사명을 입력하세요. 쉼표 또는 줄바꿈으로 구분할 수 있습니다: ")])
    if not company_names:
        raise ValueError("조회할 회사명이 필요합니다.")
    if not args.api_key:
        raise ValueError("OpenDART API 키가 필요합니다.")

    output_path = Path(args.output)
    analysis = evaluate_companies(args.api_key, company_names, args.start_year, args.end_year, output_path)
    print_results(analysis, output_path)


if __name__ == "__main__":
    main()
